# testing_app/testing_service.py

import os
from openpyxl import load_workbook, Workbook

# Excel file used ONLY by testing app
TESTING_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Notices_Testing.xlsx")

MASTER_SHEET = "Master"
GENERATED_SHEET = "2.5 Flash"  # Gemini 2.5 Flash
GPT_5_1_SHEET = "GPT 5.1"
GPT_5_MINI_SHEET = "GPT 5-Mini"

COLUMNS = [
    "PDF Name",
    "Vendor Account Number",
    "Vendor Name",
    "Service Address",
    "Notice Category",
    "Notice Date",
    "Impact Date",
    "Impact Amount",
]

def normalize_for_comparison(value: str, field_name: str = "") -> str:
    """
    Normalize field values for flexible comparison.
    - Case insensitive
    - Ignores extra spaces, commas, dashes
    - For Impact Amount: extracts only numbers
    """
    if not value or value is None:
        return ""
    
    value_str = str(value).strip()
    
    # Special handling for Impact Amount - extract only numbers
    if field_name == "Impact Amount":
        # Remove dollar signs, commas, and extract just the numeric value
        import re
        # Extract numbers including decimal points
        numbers = re.findall(r'\d+\.?\d*', value_str)
        if numbers:
            return numbers[0]  # Return first numeric value found
        return ""
    
    # For all other fields:
    # 1. Convert to lowercase (case insensitive)
    # 2. Remove extra spaces
    # 3. Remove commas
    # 4. Normalize dashes
    value_normalized = value_str.lower()
    value_normalized = ' '.join(value_normalized.split())  # Normalize whitespace
    value_normalized = value_normalized.replace(',', '')   # Remove commas
    value_normalized = value_normalized.replace('-', '')   # Remove dashes
    
    return value_normalized



def _ensure_workbook_and_sheets():
    """
    Ensure testing workbook and sheets exist with correct headers.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        wb = Workbook()
        master_ws = wb.active
        master_ws.title = MASTER_SHEET

        gen_ws = wb.create_sheet(GENERATED_SHEET)
        gen_ws.append(COLUMNS)

        gpt_5_1_ws = wb.create_sheet(GPT_5_1_SHEET)
        gpt_5_1_ws.append(COLUMNS)

        gpt_5_mini_ws = wb.create_sheet(GPT_5_MINI_SHEET)
        gpt_5_mini_ws.append(COLUMNS)

        wb.save(TESTING_EXCEL_PATH)
        return

    wb = load_workbook(TESTING_EXCEL_PATH)

    # Gemini sheet
    if GENERATED_SHEET not in wb.sheetnames:
        gen_ws = wb.create_sheet(GENERATED_SHEET)
        gen_ws.append(COLUMNS)

    # GPT 5.1 sheet
    if GPT_5_1_SHEET not in wb.sheetnames:
        gpt_5_1_ws = wb.create_sheet(GPT_5_1_SHEET)
        gpt_5_1_ws.append(COLUMNS)

    # GPT 5-mini sheet
    if GPT_5_MINI_SHEET not in wb.sheetnames:
        gpt_5_mini_ws = wb.create_sheet(GPT_5_MINI_SHEET)
        gpt_5_mini_ws.append(COLUMNS)

    wb.save(TESTING_EXCEL_PATH)


def is_pdf_in_master(pdf_name: str) -> bool:
    """
    Return True if the PDF already exists in the 'master' sheet.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return False

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return False

    ws = wb[MASTER_SHEET]
    header = [cell.value for cell in ws[1]]
    if "PDF Name" not in header:
        return False

    pdf_col = header.index("PDF Name")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cell_val = row[pdf_col]
        if cell_val and str(cell_val).strip() == pdf_name:
            return True

    return False


def _write_row_to_sheet(sheet_name: str, pdf_name: str, fields: dict):
    """
    Internal helper: upsert one row for a given sheet name.
    """
    _ensure_workbook_and_sheets()
    wb = load_workbook(TESTING_EXCEL_PATH)
    ws = wb[sheet_name]

    header_row = [cell.value for cell in ws[1]]
    if header_row != COLUMNS:
        ws.delete_rows(1, ws.max_row)
        ws.append(COLUMNS)
        header_row = COLUMNS

    col_index = {name: idx for idx, name in enumerate(header_row)}

    target_row = None
    pdf_col = col_index["PDF Name"]

    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=pdf_col + 1).value
        if cell_val and str(cell_val).strip() == pdf_name:
            target_row = row
            break

    if target_row is None:
        target_row = ws.max_row + 1

    for col_name in COLUMNS:
        col_pos = col_index[col_name]
        if col_name == "PDF Name":
            value = pdf_name
        else:
            value = fields.get(col_name, "")
        ws.cell(row=target_row, column=col_pos + 1, value=value)

    wb.save(TESTING_EXCEL_PATH)


def write_generated_row(pdf_name: str, fields: dict):
    """
    Upsert one row in 'generated' sheet for the given PDF Name. (Gemini)
    """
    _write_row_to_sheet(GENERATED_SHEET, pdf_name, fields)


def write_generated_row_gpt_5_1(pdf_name: str, fields: dict):
    """
    Upsert one row in 'GPT 5.1' sheet for the given PDF Name.
    """
    _write_row_to_sheet(GPT_5_1_SHEET, pdf_name, fields)


def write_generated_row_gpt_5_mini(pdf_name: str, fields: dict):
    """
    Upsert one row in 'GPT 5-mini' sheet for the given PDF Name.
    """
    _write_row_to_sheet(GPT_5_MINI_SHEET, pdf_name, fields)


def is_pdf_in_generated(pdf_name: str, model: str = "gemini") -> bool:
    """
    Return True if the PDF already exists in the generated sheet
    for the specified model.
    model: "gemini" | "gpt-5.1" | "gpt-5-mini"
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return False

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)

    if model == "gemini":
        sheet_name = GENERATED_SHEET
    elif model == "gpt-5.1":
        sheet_name = GPT_5_1_SHEET
    elif model == "gpt-5-mini":
        sheet_name = GPT_5_MINI_SHEET
    else:
        return False

    if sheet_name not in wb.sheetnames:
        return False

    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    if "PDF Name" not in header:
        return False

    pdf_col = header.index("PDF Name")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cell_val = row[pdf_col]
        if cell_val and str(cell_val).strip() == pdf_name:
            return True

    return False


def _sheet_to_dict_by_pdf(ws):
    header = [cell.value for cell in ws[1]]
    col_idx = {name: idx for idx, name in enumerate(header)}
    data = {}

    if "PDF Name" not in col_idx:
        return data

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        pdf_cell = row[col_idx["PDF Name"]]
        if not pdf_cell:
            continue
        pdf_name = str(pdf_cell).strip()
        fields = {}
        for col in COLUMNS:
            idx = col_idx.get(col)
            val = row[idx] if idx is not None and idx < len(row) else ""
            fields[col] = "" if val is None else str(val).strip()
        data[pdf_name] = fields
    return data


def load_and_compare_testing_excel():
    """
    Compare master vs generated for all models and return list for UI.
    Models: Master, Gemini 2.5-Flash, GPT 5.1, GPT 5-mini.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return []

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    # Require at least Master + Gemini; GPT sheets are optional
    if MASTER_SHEET not in wb.sheetnames or GENERATED_SHEET not in wb.sheetnames:
        return []

    master_ws = wb[MASTER_SHEET]
    gemini_ws = wb[GENERATED_SHEET]
    gpt_5_1_ws = wb[GPT_5_1_SHEET] if GPT_5_1_SHEET in wb.sheetnames else None
    gpt_5_mini_ws = wb[GPT_5_MINI_SHEET] if GPT_5_MINI_SHEET in wb.sheetnames else None

    master_data = _sheet_to_dict_by_pdf(master_ws)
    gemini_data = _sheet_to_dict_by_pdf(gemini_ws)
    gpt_5_1_data = _sheet_to_dict_by_pdf(gpt_5_1_ws) if gpt_5_1_ws else {}
    gpt_5_mini_data = _sheet_to_dict_by_pdf(gpt_5_mini_ws) if gpt_5_mini_ws else {}

    # Only compare PDFs that exist in master
    all_pdfs = sorted(master_data.keys())
    results = []
    s_no = 1

    for pdf_name in all_pdfs:
        master_fields = master_data.get(pdf_name)
        if not master_fields:
            continue

        # Per-model fields (may be missing)
        gemini_fields = gemini_data.get(pdf_name)
        gpt_5_1_fields = gpt_5_1_data.get(pdf_name)
        gpt_5_mini_fields = gpt_5_mini_data.get(pdf_name)

        # Function to compute mismatches vs master
        def compare_fields(model_fields: dict | None) -> tuple[dict, str]:
            if not model_fields:
                return {col: True for col in COLUMNS[1:]}, "missing"
            
            mismatches = {}
            for col in COLUMNS[1:]:
                m_val = master_fields.get(col, "")
                g_val = model_fields.get(col, "")
                # Normalize both values before comparison
                m_normalized = normalize_for_comparison(m_val, col)
                g_normalized = normalize_for_comparison(g_val, col)
                mismatches[col] = (m_normalized != g_normalized)
            
            # Check ONLY vendor fields for "Vendor Identified" status
            vendor_fields = ["Vendor Account Number", "Vendor Name", "Service Address"]
            vendor_match = all(not mismatches.get(f, True) for f in vendor_fields)
            
            if vendor_match:
                return mismatches, "correct"
            return mismatches, "incorrect"


        gemini_mismatches, gemini_status = compare_fields(gemini_fields)
        gpt_5_1_mismatches, gpt_5_1_status = compare_fields(gpt_5_1_fields)
        gpt_5_mini_mismatches, gpt_5_mini_status = compare_fields(gpt_5_mini_fields)

        results.append(
            {
                "s_no": s_no,
                "pdf_name": pdf_name,
                "rows": [
                    {
                        "model": "Master",
                        "vendor_account": master_fields["Vendor Account Number"],
                        "vendor_name": master_fields["Vendor Name"],
                        "service_address": master_fields["Service Address"],
                        "category": master_fields["Notice Category"],
                        "notice_date": master_fields["Notice Date"],
                        "impact_date": master_fields["Impact Date"],
                        "impact_amount": master_fields["Impact Amount"],
                        "details_verified": None,
                        "field_mismatches": {},
                    },
                    {
                        "model": "Gemini 2.5-Flash",
                        "vendor_account": (gemini_fields or {}).get(
                            "Vendor Account Number", ""
                        ),
                        "vendor_name": (gemini_fields or {}).get("Vendor Name", ""),
                        "service_address": (gemini_fields or {}).get(
                            "Service Address", ""
                        ),
                        "category": (gemini_fields or {}).get("Notice Category", ""),
                        "notice_date": (gemini_fields or {}).get("Notice Date", ""),
                        "impact_date": (gemini_fields or {}).get("Impact Date", ""),
                        "impact_amount": (gemini_fields or {}).get("Impact Amount", ""),
                        "details_verified": gemini_status,
                        "field_mismatches": gemini_mismatches,
                    },
                    {
                        "model": "GPT 5.1",
                        "vendor_account": (gpt_5_1_fields or {}).get(
                            "Vendor Account Number", ""
                        ),
                        "vendor_name": (gpt_5_1_fields or {}).get("Vendor Name", ""),
                        "service_address": (gpt_5_1_fields or {}).get(
                            "Service Address", ""
                        ),
                        "category": (gpt_5_1_fields or {}).get("Notice Category", ""),
                        "notice_date": (gpt_5_1_fields or {}).get("Notice Date", ""),
                        "impact_date": (gpt_5_1_fields or {}).get("Impact Date", ""),
                        "impact_amount": (gpt_5_1_fields or {}).get("Impact Amount", ""),
                        "details_verified": gpt_5_1_status,
                        "field_mismatches": gpt_5_1_mismatches,
                    },
                    {
                        "model": "GPT 5-mini",
                        "vendor_account": (gpt_5_mini_fields or {}).get(
                            "Vendor Account Number", ""
                        ),
                        "vendor_name": (gpt_5_mini_fields or {}).get("Vendor Name", ""),
                        "service_address": (gpt_5_mini_fields or {}).get(
                            "Service Address", ""
                        ),
                        "category": (gpt_5_mini_fields or {}).get(
                            "Notice Category", ""
                        ),
                        "notice_date": (gpt_5_mini_fields or {}).get("Notice Date", ""),
                        "impact_date": (gpt_5_mini_fields or {}).get("Impact Date", ""),
                        "impact_amount": (gpt_5_mini_fields or {}).get(
                            "Impact Amount", ""
                        ),
                        "details_verified": gpt_5_mini_status,
                        "field_mismatches": gpt_5_mini_mismatches,
                    },
                ],
            }
        )
        s_no += 1

    return results

def load_and_compare_gemini_only():
    """
    Compare Master vs Gemini only.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return []

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames or GENERATED_SHEET not in wb.sheetnames:
        return []

    master_ws = wb[MASTER_SHEET]
    gemini_ws = wb[GENERATED_SHEET]

    master_data = _sheet_to_dict_by_pdf(master_ws)
    gemini_data = _sheet_to_dict_by_pdf(gemini_ws)

    all_pdfs = sorted(master_data.keys())
    results = []
    s_no = 1

    for pdf_name in all_pdfs:
        master_fields = master_data.get(pdf_name)
        gemini_fields = gemini_data.get(pdf_name)
        if not master_fields or not gemini_fields:
            continue

        field_mismatches = {}
        for col in COLUMNS[1:]:
            m_val = master_fields.get(col, "")
            g_val = gemini_fields.get(col, "")
            m_normalized = normalize_for_comparison(m_val, col)
            g_normalized = normalize_for_comparison(g_val, col)
            field_mismatches[col] = (m_normalized != g_normalized)


        # Check ONLY vendor fields
        vendor_fields = ["Vendor Account Number", "Vendor Name", "Service Address"]
        vendor_match = all(not field_mismatches.get(f, True) for f in vendor_fields)
        details_verified = "correct" if vendor_match else "incorrect"


        results.append(
            {
                "s_no": s_no,
                "pdf_name": pdf_name,
                "rows": [
                    {
                        "model": "Master",
                        "vendor_account": master_fields["Vendor Account Number"],
                        "vendor_name": master_fields["Vendor Name"],
                        "service_address": master_fields["Service Address"],
                        "category": master_fields["Notice Category"],
                        "notice_date": master_fields["Notice Date"],
                        "impact_date": master_fields["Impact Date"],
                        "impact_amount": master_fields["Impact Amount"],
                        "details_verified": None,
                        "field_mismatches": {},
                    },
                    {
                        "model": "Gemini 2.5-Flash",
                        "vendor_account": gemini_fields["Vendor Account Number"],
                        "vendor_name": gemini_fields["Vendor Name"],
                        "service_address": gemini_fields["Service Address"],
                        "category": gemini_fields["Notice Category"],
                        "notice_date": gemini_fields["Notice Date"],
                        "impact_date": gemini_fields["Impact Date"],
                        "impact_amount": gemini_fields["Impact Amount"],
                        "details_verified": details_verified,
                        "field_mismatches": field_mismatches,
                    },
                ],
            }
        )
        s_no += 1

    return results


def load_and_compare_gpt_5_1_only():
    """
    Compare Master vs GPT 5.1 only.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return []

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames or GPT_5_1_SHEET not in wb.sheetnames:
        return []

    master_ws = wb[MASTER_SHEET]
    gpt_ws = wb[GPT_5_1_SHEET]

    master_data = _sheet_to_dict_by_pdf(master_ws)
    gpt_data = _sheet_to_dict_by_pdf(gpt_ws)

    all_pdfs = sorted(master_data.keys())
    results = []
    s_no = 1

    for pdf_name in all_pdfs:
        master_fields = master_data.get(pdf_name)
        gpt_fields = gpt_data.get(pdf_name)
        if not master_fields or not gpt_fields:
            continue

        field_mismatches = {}
        for col in COLUMNS[1:]:
            m_val = master_fields.get(col, "")
            g_val = gpt_fields.get(col, "")
            m_normalized = normalize_for_comparison(m_val, col)
            g_normalized = normalize_for_comparison(g_val, col)
            field_mismatches[col] = (m_normalized != g_normalized)


        # Check ONLY vendor fields
        vendor_fields = ["Vendor Account Number", "Vendor Name", "Service Address"]
        vendor_match = all(not field_mismatches.get(f, True) for f in vendor_fields)
        details_verified = "correct" if vendor_match else "incorrect"


        results.append(
            {
                "s_no": s_no,
                "pdf_name": pdf_name,
                "rows": [
                    {
                        "model": "Master",
                        "vendor_account": master_fields["Vendor Account Number"],
                        "vendor_name": master_fields["Vendor Name"],
                        "service_address": master_fields["Service Address"],
                        "category": master_fields["Notice Category"],
                        "notice_date": master_fields["Notice Date"],
                        "impact_date": master_fields["Impact Date"],
                        "impact_amount": master_fields["Impact Amount"],
                        "details_verified": None,
                        "field_mismatches": {},
                    },
                    {
                        "model": "GPT 5.1",
                        "vendor_account": gpt_fields["Vendor Account Number"],
                        "vendor_name": gpt_fields["Vendor Name"],
                        "service_address": gpt_fields["Service Address"],
                        "category": gpt_fields["Notice Category"],
                        "notice_date": gpt_fields["Notice Date"],
                        "impact_date": gpt_fields["Impact Date"],
                        "impact_amount": gpt_fields["Impact Amount"],
                        "details_verified": details_verified,
                        "field_mismatches": field_mismatches,
                    },
                ],
            }
        )
        s_no += 1

    return results


def load_and_compare_gpt_5_mini_only():
    """
    Compare Master vs GPT 5-mini only.
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return []

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames or GPT_5_MINI_SHEET not in wb.sheetnames:
        return []

    master_ws = wb[MASTER_SHEET]
    gpt_ws = wb[GPT_5_MINI_SHEET]

    master_data = _sheet_to_dict_by_pdf(master_ws)
    gpt_data = _sheet_to_dict_by_pdf(gpt_ws)

    all_pdfs = sorted(master_data.keys())
    results = []
    s_no = 1

    for pdf_name in all_pdfs:
        master_fields = master_data.get(pdf_name)
        gpt_fields = gpt_data.get(pdf_name)
        if not master_fields or not gpt_fields:
            continue

        field_mismatches = {}
        for col in COLUMNS[1:]:
            m_val = master_fields.get(col, "")
            g_val = gpt_fields.get(col, "")
            m_normalized = normalize_for_comparison(m_val, col)
            g_normalized = normalize_for_comparison(g_val, col)
            field_mismatches[col] = (m_normalized != g_normalized)


        # Check ONLY vendor fields
        vendor_fields = ["Vendor Account Number", "Vendor Name", "Service Address"]
        vendor_match = all(not field_mismatches.get(f, True) for f in vendor_fields)
        details_verified = "correct" if vendor_match else "incorrect"


        results.append(
            {
                "s_no": s_no,
                "pdf_name": pdf_name,
                "rows": [
                    {
                        "model": "Master",
                        "vendor_account": master_fields["Vendor Account Number"],
                        "vendor_name": master_fields["Vendor Name"],
                        "service_address": master_fields["Service Address"],
                        "category": master_fields["Notice Category"],
                        "notice_date": master_fields["Notice Date"],
                        "impact_date": master_fields["Impact Date"],
                        "impact_amount": master_fields["Impact Amount"],
                        "details_verified": None,
                        "field_mismatches": {},
                    },
                    {
                        "model": "GPT 5-mini",
                        "vendor_account": gpt_fields["Vendor Account Number"],
                        "vendor_name": gpt_fields["Vendor Name"],
                        "service_address": gpt_fields["Service Address"],
                        "category": gpt_fields["Notice Category"],
                        "notice_date": gpt_fields["Notice Date"],
                        "impact_date": gpt_fields["Impact Date"],
                        "impact_amount": gpt_fields["Impact Amount"],
                        "details_verified": details_verified,
                        "field_mismatches": field_mismatches,
                    },
                ],
            }
        )
        s_no += 1

    return results

def calculate_overall_accuracy():
    """
    Calculate overall vendor identification accuracy for each model.
    Returns: {
        "gemini": 88.5,
        "gpt_5_1": 92.3,
        "gpt_5_mini": 80.0
    }
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    master_ws = wb[MASTER_SHEET]
    master_data = _sheet_to_dict_by_pdf(master_ws)
    
    total_pdfs = len(master_data)
    if total_pdfs == 0:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    vendor_fields = ["Vendor Account Number", "Vendor Name", "Service Address"]

    def count_correct(sheet_name):
        if sheet_name not in wb.sheetnames:
            return 0
        
        model_ws = wb[sheet_name]
        model_data = _sheet_to_dict_by_pdf(model_ws)
        
        correct_count = 0
        for pdf_name, master_fields in master_data.items():
            model_fields = model_data.get(pdf_name)
            if not model_fields:
                continue
            
            # Check if vendor fields match
            vendor_match = all(
                normalize_for_comparison(master_fields.get(field, ""), field) == 
                normalize_for_comparison(model_fields.get(field, ""), field)
                for field in vendor_fields
            )

            
            if vendor_match:
                correct_count += 1
        
        return correct_count

    gemini_correct = count_correct(GENERATED_SHEET)
    gpt_5_1_correct = count_correct(GPT_5_1_SHEET)
    gpt_5_mini_correct = count_correct(GPT_5_MINI_SHEET)

    return {
        "gemini": round((gemini_correct / total_pdfs) * 100, 1),
        "gpt_5_1": round((gpt_5_1_correct / total_pdfs) * 100, 1),
        "gpt_5_mini": round((gpt_5_mini_correct / total_pdfs) * 100, 1)
    }


def calculate_category_accuracy():
    """
    Calculate category-wise classification accuracy percentages for each model.
    Returns: {
        "categories": ["Late Notice", "Maintenance", ...],
        "gemini_accuracy": [100.0, 85.5, 66.7, ...],
        "gpt_5_1_accuracy": [95.0, 100.0, 75.0, ...],
        "gpt_5_mini_accuracy": [90.0, 80.0, 83.3, ...]
    }
    """
    categories = [
        'Late Notice',
        'Maintenance',
        'Address Change',
        'Cheque Received',
        'Disconnect Notice',
        'Rate Change',
        'Revert to Owner',
        '3rd Party Audit',
        'Others'
    ]
    
    if not os.path.exists(TESTING_EXCEL_PATH):
        return {
            "categories": categories,
            "gemini_accuracy": [0] * 9,
            "gpt_5_1_accuracy": [0] * 9,
            "gpt_5_mini_accuracy": [0] * 9
        }

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return {
            "categories": categories,
            "gemini_accuracy": [0] * 9,
            "gpt_5_1_accuracy": [0] * 9,
            "gpt_5_mini_accuracy": [0] * 9
        }

    master_ws = wb[MASTER_SHEET]
    master_data = _sheet_to_dict_by_pdf(master_ws)

    # First, count total documents per category
    category_totals = {cat: 0 for cat in categories}
    for pdf_name, master_fields in master_data.items():
        master_category = master_fields.get('Notice Category', 'Others')
        if master_category in category_totals:
            category_totals[master_category] += 1
        else:
            category_totals['Others'] += 1

    def calculate_accuracy_per_category(sheet_name):
        if sheet_name not in wb.sheetnames:
            return [0] * 9
        
        model_ws = wb[sheet_name]
        model_data = _sheet_to_dict_by_pdf(model_ws)
        
        category_correct = {cat: 0 for cat in categories}
        
        for pdf_name, master_fields in master_data.items():
            model_fields = model_data.get(pdf_name)
            if not model_fields:
                continue
            
            master_category = master_fields.get('Notice Category', 'Others')
            model_category = model_fields.get('Notice Category', 'Others')
            
            # Check if the model correctly classified the category
            if master_category == model_category:
                if master_category in category_correct:
                    category_correct[master_category] += 1
                else:
                    category_correct['Others'] += 1
        
        # Calculate accuracy percentage for each category
        accuracy_percentages = []
        for cat in categories:
            total = category_totals[cat]
            correct = category_correct[cat]
            if total > 0:
                accuracy = (correct / total) * 100
                accuracy_percentages.append(round(accuracy, 1))
            else:
                accuracy_percentages.append(0)
        
        return accuracy_percentages

    gemini_accuracy = calculate_accuracy_per_category(GENERATED_SHEET)
    gpt_5_1_accuracy = calculate_accuracy_per_category(GPT_5_1_SHEET)
    gpt_5_mini_accuracy = calculate_accuracy_per_category(GPT_5_MINI_SHEET)

    return {
        "categories": categories,
        "gemini_accuracy": gemini_accuracy,
        "gpt_5_1_accuracy": gpt_5_1_accuracy,
        "gpt_5_mini_accuracy": gpt_5_mini_accuracy
    }

def calculate_disconnect_late_accuracy():
    """
    Calculate Impact Date and Impact Amount accuracy for Disconnect and Late Notice only.
    Returns: {
        "gemini": 88.5,
        "gpt_5_1": 92.3,
        "gpt_5_mini": 80.0
    }
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    master_ws = wb[MASTER_SHEET]
    master_data = _sheet_to_dict_by_pdf(master_ws)
    
    # Filter only Disconnect and Late Notice
    relevant_pdfs = {
        pdf: fields for pdf, fields in master_data.items()
        if fields.get('Notice Category') in ['Disconnect Notice', 'Late Notice']
    }
    
    total_pdfs = len(relevant_pdfs)
    if total_pdfs == 0:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    fields_to_check = [ "Impact Amount"]

    def count_correct(sheet_name):
        if sheet_name not in wb.sheetnames:
            return 0
        
        model_ws = wb[sheet_name]
        model_data = _sheet_to_dict_by_pdf(model_ws)
        
        correct_count = 0
        for pdf_name, master_fields in relevant_pdfs.items():
            model_fields = model_data.get(pdf_name)
            if not model_fields:
                continue
            
            # Check if both Impact Date AND Impact Amount match
            all_match = all(
                master_fields.get(field, "") == model_fields.get(field, "")
                for field in fields_to_check
            )
            
            if all_match:
                correct_count += 1
        
        return correct_count

    gemini_correct = count_correct(GENERATED_SHEET)
    gpt_5_1_correct = count_correct(GPT_5_1_SHEET)
    gpt_5_mini_correct = count_correct(GPT_5_MINI_SHEET)

    return {
        "gemini": round((gemini_correct / total_pdfs) * 100, 1),
        "gpt_5_1": round((gpt_5_1_correct / total_pdfs) * 100, 1),
        "gpt_5_mini": round((gpt_5_mini_correct / total_pdfs) * 100, 1)
    }

def calculate_date_accuracy():
    """
    Calculate Impact Date and Impact Amount accuracy for Disconnect and Late Notice only.
    Returns: {
        "gemini": 88.5,
        "gpt_5_1": 92.3,
        "gpt_5_mini": 80.0
    }
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    master_ws = wb[MASTER_SHEET]
    master_data = _sheet_to_dict_by_pdf(master_ws)
    
    # Filter only Disconnect and Late Notice
    relevant_pdfs = {
        pdf: fields for pdf, fields in master_data.items()
        if fields.get('Notice Category') in ['Disconnect Notice', 'Late Notice']
    }
    
    total_pdfs = len(relevant_pdfs)
    if total_pdfs == 0:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    fields_to_check = [ "Impact Date"]

    def count_correct(sheet_name):
        if sheet_name not in wb.sheetnames:
            return 0
        
        model_ws = wb[sheet_name]
        model_data = _sheet_to_dict_by_pdf(model_ws)
        
        correct_count = 0
        for pdf_name, master_fields in relevant_pdfs.items():
            model_fields = model_data.get(pdf_name)
            if not model_fields:
                continue
            
            # Check if both Impact Date AND Impact Amount match
            all_match = all(
                master_fields.get(field, "") == model_fields.get(field, "")
                for field in fields_to_check
            )
            
            if all_match:
                correct_count += 1
        
        return correct_count

    gemini_correct = count_correct(GENERATED_SHEET)
    gpt_5_1_correct = count_correct(GPT_5_1_SHEET)
    gpt_5_mini_correct = count_correct(GPT_5_MINI_SHEET)

    return {
        "gemini": round((gemini_correct / total_pdfs) * 100, 1),
        "gpt_5_1": round((gpt_5_1_correct / total_pdfs) * 100, 1),
        "gpt_5_mini": round((gpt_5_mini_correct / total_pdfs) * 100, 1)
    }

def calculate_notice_date_accuracy():
    """
    Calculate Notice Date accuracy for all notice types.
    Returns: {
        "gemini": 88.5,
        "gpt_5_1": 92.3,
        "gpt_5_mini": 80.0
    }
    """
    if not os.path.exists(TESTING_EXCEL_PATH):
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    wb = load_workbook(TESTING_EXCEL_PATH, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    master_ws = wb[MASTER_SHEET]
    master_data = _sheet_to_dict_by_pdf(master_ws)
    
    total_pdfs = len(master_data)
    if total_pdfs == 0:
        return {"gemini": 0, "gpt_5_1": 0, "gpt_5_mini": 0}

    def count_correct(sheet_name):
        if sheet_name not in wb.sheetnames:
            return 0
        
        model_ws = wb[sheet_name]
        model_data = _sheet_to_dict_by_pdf(model_ws)
        
        correct_count = 0
        for pdf_name, master_fields in master_data.items():
            model_fields = model_data.get(pdf_name)
            if not model_fields:
                continue
            
            # Check if Notice Date matches
            if master_fields.get("Notice Date", "") == model_fields.get("Notice Date", ""):
                correct_count += 1
        
        return correct_count

    gemini_correct = count_correct(GENERATED_SHEET)
    gpt_5_1_correct = count_correct(GPT_5_1_SHEET)
    gpt_5_mini_correct = count_correct(GPT_5_MINI_SHEET)

    return {
        "gemini": round((gemini_correct / total_pdfs) * 100, 1),
        "gpt_5_1": round((gpt_5_1_correct / total_pdfs) * 100, 1),
        "gpt_5_mini": round((gpt_5_mini_correct / total_pdfs) * 100, 1)
    }
def calculate_correct_row_counts():
    """
    Counts rows where ALL required fields match the master row.
    """

    comparison = load_and_compare_testing_excel()

    result = {
        "gemini": {"correct": 0, "total": 0},
        "gpt_5_1": {"correct": 0, "total": 0},
        "gpt_5_mini": {"correct": 0, "total": 0},
    }

    REQUIRED_FIELDS = {
        "Vendor Account Number",
        "Vendor Name",
        "Service Address",
        "Notice Category",
        "Notice Date",
        "Impact Date",
        "Impact Amount",
    }

    for item in comparison:
        rows = item["rows"]

        # rows[0] = master row
        for row in rows[1:]:  # model rows only
            raw_model = row["model"]

            if "gemini" in raw_model.lower():
                model_key = "gemini"
            elif "gpt 5.1" in raw_model.lower():
                model_key = "gpt_5_1"
            elif "gpt 5-mini" in raw_model.lower() or "gpt 5 mini" in raw_model.lower():
                model_key = "gpt_5_mini"
            else:
                continue


            result[model_key]["total"] += 1


            mismatches = row.get("field_mismatches", {})

            # Perfect row = NO mismatch in any required field
            is_perfect = True
            for field in REQUIRED_FIELDS:
                if mismatches.get(field) is True:
                    is_perfect = False
                    break

            if is_perfect:
                result[model_key]["correct"] += 1

    print("MODEL NAME FROM EXCEL:", row["model"])

    return result
